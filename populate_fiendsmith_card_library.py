#!/usr/bin/env python3
import argparse
import json
import re
import time
from datetime import date
from pathlib import Path

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Alignment

BASE_URL = "https://www.db.yugioh-card.com"
SEARCH_PATH = "/yugiohdb/card_search.action"

ATTRIBUTES = {"LIGHT", "DARK", "EARTH", "FIRE", "WATER", "WIND", "DIVINE"}
EXTRA_DECK_TYPES = {"Fusion", "Synchro", "Xyz", "Link"}

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
}


def normalize_name(name: str) -> str:
    name = name.strip()
    name = name.replace("\u2019", "'")
    name = name.replace("\u2010", "-")
    name = name.replace("\u2011", "-")
    name = name.replace("\u2013", "-")
    return name


def simplify_name(name: str) -> str:
    return re.sub(r"[^A-Za-z0-9]+", "", normalize_name(name)).lower()


def ensure_cache_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def build_cache_index(cache_dir: Path) -> dict[str, list[dict[str, str]]]:
    index: dict[str, list[dict[str, str]]] = {}
    for path in cache_dir.glob("card_*.html"):
        html = path.read_text(encoding="utf-8")
        details = parse_card_details(html)
        name = details.get("name", "").strip()
        if not name:
            continue
        cid = path.stem.replace("card_", "")
        url = f"{BASE_URL}/yugiohdb/card_search.action?ope=2&cid={cid}&request_locale=en"
        key = normalize_name(name).lower()
        index.setdefault(key, []).append({"cid": cid, "name": name, "url": url})
    return index


def search_cards(session: requests.Session, name: str) -> list[tuple[str, str]]:
    params = {"ope": "1", "keyword": name, "request_locale": "en"}
    resp = session.get(BASE_URL + SEARCH_PATH, params=params, headers=HEADERS, timeout=30)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "lxml")
    results = []
    for row in soup.select("div.t_row"):
        cnm = row.select_one("input.cnm")
        link = row.select_one("input.link_value")
        if cnm and link:
            results.append((cnm["value"], link["value"]))
    return results


def pick_best_match(name: str, results: list[tuple[str, str]]) -> tuple[str, str] | None:
    norm = normalize_name(name).lower()
    exact = [r for r in results if normalize_name(r[0]).lower() == norm]
    if len(exact) == 1:
        return exact[0]
    if len(exact) > 1:
        return None

    simp = simplify_name(name)
    simp_matches = [r for r in results if simplify_name(r[0]) == simp]
    if len(simp_matches) == 1:
        return simp_matches[0]

    if len(results) == 1:
        return results[0]

    return None


def build_detail_url(link_value: str) -> str:
    if link_value.startswith("http"):
        url = link_value
    else:
        url = BASE_URL + link_value
    if "request_locale=" not in url:
        url += "&request_locale=en" if "?" in url else "?request_locale=en"
    return url


def extract_cid(url: str) -> str | None:
    match = re.search(r"cid=(\d+)", url)
    return match.group(1) if match else None


def load_card_html(
    session: requests.Session,
    url: str,
    cache_dir: Path,
    offline: bool,
    rate_limit_seconds: float,
) -> str:
    cid = extract_cid(url)
    cache_path = cache_dir / f"card_{cid}.html" if cid else None
    if cache_path and cache_path.exists():
        return cache_path.read_text(encoding="utf-8")

    if offline:
        raise FileNotFoundError(f"Offline mode: missing cached HTML for {url}")

    resp = session.get(url, headers=HEADERS, timeout=30)
    resp.raise_for_status()
    html = resp.text
    if cache_path:
        cache_path.write_text(html, encoding="utf-8")
    if rate_limit_seconds > 0:
        time.sleep(rate_limit_seconds)
    return html


def parse_card_details(html: str) -> dict:
    soup = BeautifulSoup(html, "lxml")
    name_tag = soup.select_one("div#cardname h1")
    name = name_tag.get_text(strip=True) if name_tag else ""

    card_text = ""
    text_block = None
    for block in soup.select("div.CardText"):
        if block.select_one(".item_box_text"):
            text_block = block
            break
    if text_block:
        item_box_text = text_block.select_one(".item_box_text")
        raw = item_box_text.get_text("\n", strip=True) if item_box_text else ""
        lines = [line.strip() for line in raw.split("\n") if line.strip()]
        if lines and lines[0].lower() == "card text":
            lines = lines[1:]
        card_text = "\n".join(lines)

    spec_block = None
    for block in soup.select("div.CardText"):
        if block.select_one(".item_box_text"):
            continue
        classes = block.get("class") or []
        if "CardLanguage" in classes:
            continue
        spec_block = block
        break

    card_type = ""
    subtype_icon = ""
    attribute = ""
    monster_typing = ""
    level_rank_link = ""
    atk = ""
    defense = ""

    if spec_block:
        icon_value = None
        for item in spec_block.select(".item_box"):
            title_tag = item.select_one(".item_box_title")
            value_tag = item.select_one(".item_box_value")
            title_text = title_tag.get_text(strip=True) if title_tag else ""
            value_text = value_tag.get_text(strip=True) if value_tag else ""

            if title_text == "Icon":
                icon_value = value_text
            elif title_text == "ATK":
                atk = value_text
            elif title_text == "DEF":
                defense = value_text
            elif value_text in ATTRIBUTES:
                attribute = value_text
            elif value_text.startswith("Level ") or value_text.startswith("Rank ") or value_text.startswith("Link "):
                level_rank_link = value_text

        species = spec_block.select_one("p.species")
        if species:
            typing_text = species.get_text(" ", strip=True)
            typing_text = typing_text.replace("\uFF0F", "/")
            typing_text = re.sub(r"\s*/\s*", " / ", typing_text)
            monster_typing = typing_text

        if icon_value:
            subtype_icon = icon_value
            if "Spell" in icon_value:
                card_type = "Spell"
            elif "Trap" in icon_value:
                card_type = "Trap"
        else:
            if attribute or monster_typing:
                card_type = "Monster"
                if monster_typing:
                    parts = [p.strip() for p in monster_typing.split("/") if p.strip()]
                    if len(parts) > 1:
                        subtype_icon = " / ".join(parts[1:])
            else:
                card_type = ""

    return {
        "name": name,
        "card_text": card_text,
        "card_type": card_type,
        "subtype_icon": subtype_icon,
        "attribute": attribute,
        "monster_typing": monster_typing,
        "level_rank_link": level_rank_link,
        "atk": atk,
        "defense": defense,
    }


def is_extra_deck(subtype_icon: str) -> bool:
    if not subtype_icon:
        return False
    for t in EXTRA_DECK_TYPES:
        if t.lower() in subtype_icon.lower():
            return True
    return False


def extract_materials(card_text: str) -> str:
    if not card_text:
        return ""
    lines = [line.strip() for line in card_text.split("\n") if line.strip()]
    if not lines:
        return ""
    if len(lines) == 1:
        return lines[0]

    def is_effect_line(line: str) -> bool:
        if re.search(r"[\.:;]", line):
            return True
        if re.search(r"\b(When|If|During|Once|You can|Cannot|Must be|Target|Tribute|Special Summon|Destroy)\b", line):
            return True
        return False

    materials = []
    for line in lines:
        if materials and is_effect_line(line):
            break
        if not materials and is_effect_line(line):
            break
        materials.append(line)

    return "\n".join(materials) if materials else ""


def parse_args() -> argparse.Namespace:
    repo_root = Path(__file__).resolve().parent
    parser = argparse.ArgumentParser(description="Populate the clean card library from Konami DB.")
    parser.add_argument("--input-xlsx", required=True, help="Path to input XLSX.")
    parser.add_argument(
        "--output-xlsx",
        default=str(repo_root / "data_processed" / "Fiendsmith_Master_Card_Library_CLEAN.xlsx"),
    )
    parser.add_argument(
        "--output-json",
        default=str(repo_root / "data_processed" / "Fiendsmith_Master_Card_Library_CLEAN.json"),
    )
    parser.add_argument(
        "--cache-dir",
        default=str(repo_root / "data_cache" / "konami_db"),
    )
    parser.add_argument("--offline", action="store_true", help="Use cached HTML only.")
    parser.add_argument("--rate-limit-seconds", type=float, default=0.7)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    input_path = Path(args.input_xlsx)
    output_xlsx = Path(args.output_xlsx)
    output_json = Path(args.output_json)
    cache_dir = Path(args.cache_dir)

    if not input_path.exists():
        raise SystemExit(f"Missing input XLSX: {input_path}")

    ensure_cache_dir(cache_dir)
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    output_json.parent.mkdir(parents=True, exist_ok=True)

    df = pd.read_excel(input_path, sheet_name="Sheet1")

    session = requests.Session()

    cache_index = {}
    if args.offline:
        cache_index = build_cache_index(cache_dir)

    today = date.today().isoformat()
    rows = []

    for _, row in df.iterrows():
        original_name = str(row.get("Name", "")).strip()
        normalized_name = normalize_name(original_name)

        lookup_notes = ""
        match = None
        if args.offline:
            key = normalize_name(normalized_name).lower()
            matches = cache_index.get(key, [])
            if len(matches) == 1:
                match = (matches[0]["name"], matches[0]["url"])
            elif len(matches) > 1:
                lookup_notes = "Ambiguous cache match in offline mode"
            else:
                raise SystemExit(f"Offline cache miss for card: {normalized_name}")
        else:
            try:
                results = search_cards(session, normalized_name)
                match = pick_best_match(normalized_name, results)
                if not match:
                    lookup_notes = "Ambiguous or no match on official DB search"
            except Exception as exc:
                lookup_notes = f"Search error: {exc}"

        if not match:
            rows.append({
                "Name (Official formatting)": normalized_name,
                "Card Type (Monster / Spell / Trap)": "",
                "Subtype/Icon": "",
                "Attribute": "",
                "Monster Typing": "",
                "Level/Rank/Link": "",
                "ATK": "",
                "DEF": "",
                "Summoning Requirements / Materials": "",
                "Official Card Text (Exact, TCG)": "NOT FOUND",
                "Official DB URL": "",
                "CID": "",
                "Verified Date": today,
                "Lookup Notes": lookup_notes,
            })
            continue

        official_name, link_value = match
        detail_url = build_detail_url(link_value)
        cid = extract_cid(detail_url) or ""

        try:
            html = load_card_html(session, detail_url, cache_dir, args.offline, args.rate_limit_seconds)
            details = parse_card_details(html)
        except Exception as exc:
            if args.offline:
                raise SystemExit(f"Offline cache miss for CID {cid}: {exc}")
            rows.append({
                "Name (Official formatting)": official_name,
                "Card Type (Monster / Spell / Trap)": "",
                "Subtype/Icon": "",
                "Attribute": "",
                "Monster Typing": "",
                "Level/Rank/Link": "",
                "ATK": "",
                "DEF": "",
                "Summoning Requirements / Materials": "",
                "Official Card Text (Exact, TCG)": "NOT FOUND",
                "Official DB URL": "",
                "CID": "",
                "Verified Date": today,
                "Lookup Notes": f"Detail fetch/parse error: {exc}",
            })
            continue

        card_text = details.get("card_text", "")
        if not card_text:
            lookup_notes = "Card Text not found on page"
            card_text = "NOT FOUND"

        subtype_icon = details.get("subtype_icon", "")
        materials = ""
        if is_extra_deck(subtype_icon) and card_text and card_text != "NOT FOUND":
            materials = extract_materials(card_text)

        card_name = details.get("name") or official_name
        card_name = normalize_name(card_name)

        rows.append({
            "Name (Official formatting)": card_name,
            "Card Type (Monster / Spell / Trap)": details.get("card_type", ""),
            "Subtype/Icon": subtype_icon,
            "Attribute": details.get("attribute", ""),
            "Monster Typing": details.get("monster_typing", ""),
            "Level/Rank/Link": details.get("level_rank_link", ""),
            "ATK": details.get("atk", ""),
            "DEF": details.get("defense", ""),
            "Summoning Requirements / Materials": materials,
            "Official Card Text (Exact, TCG)": card_text,
            "Official DB URL": detail_url if card_text != "NOT FOUND" else "",
            "CID": cid if card_text != "NOT FOUND" else "",
            "Verified Date": today,
            "Lookup Notes": lookup_notes,
        })

    output_columns = [
        "Name (Official formatting)",
        "Card Type (Monster / Spell / Trap)",
        "Subtype/Icon",
        "Attribute",
        "Monster Typing",
        "Level/Rank/Link",
        "ATK",
        "DEF",
        "Summoning Requirements / Materials",
        "Official Card Text (Exact, TCG)",
        "Official DB URL",
        "CID",
        "Verified Date",
        "Lookup Notes",
    ]

    out_df = pd.DataFrame(rows, columns=output_columns)

    # QA checks
    name_series = out_df["Name (Official formatting)"].fillna("")
    if name_series.duplicated().any():
        raise SystemExit("Duplicate card names detected after normalization.")
    if (out_df["Official Card Text (Exact, TCG)"].fillna("") == "").any():
        raise SystemExit("Blank Official Card Text detected.")
    if len(out_df) != len(df):
        raise SystemExit("Row count mismatch after processing.")

    out_df.to_excel(output_xlsx, index=False)

    wb = load_workbook(output_xlsx)
    ws = wb.active
    ws.freeze_panes = "A2"

    # Wrap text in Official Card Text column (J)
    for cell in ws["J"]:
        cell.alignment = Alignment(wrap_text=True)

    # Widen Official Card Text column
    ws.column_dimensions["J"].width = 80

    wb.save(output_xlsx)

    with output_json.open("w", encoding="utf-8") as f:
        for _, record in out_df.iterrows():
            obj = {col: ("" if pd.isna(record[col]) else record[col]) for col in output_columns}
            f.write(json.dumps(obj, ensure_ascii=True) + "\n")


if __name__ == "__main__":
    main()
